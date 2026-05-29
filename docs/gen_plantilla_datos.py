#!/usr/bin/env python3
"""Genera el Excel-plantilla de datos Fase 1 para el cliente (Agua Viva).
Mapea 1:1 a los modelos OpenEduCat/SciBack instalados. Salida: plantilla-datos-colegio.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

AZUL = "1F4E78"; AZUL_CLARO = "D6E4F0"; AMBAR = "FCE4D6"; GRIS = "808080"
hdr_fill = PatternFill("solid", fgColor=AZUL)
hdr_font = Font(color="FFFFFF", bold=True, size=10)
oblig_fill = PatternFill("solid", fgColor=AMBAR)   # columnas obligatorias
opt_fill = PatternFill("solid", fgColor=AZUL_CLARO)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

def hoja(nombre, titulo, cols):
    """cols: lista de (encabezado, obligatorio_bool, ancho, ayuda, ejemplo)"""
    ws = wb.create_sheet(nombre)
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    t = ws.cell(1, 1, titulo); t.font = Font(bold=True, size=13, color=AZUL); t.alignment = left
    leyenda = ws.cell(2, 1, "Naranja = OBLIGATORIO   ·   Azul = opcional   ·   no borre los encabezados")
    leyenda.font = Font(italic=True, size=9, color=GRIS)
    for c, (h, oblig, w, ayuda, ej) in enumerate(cols, start=1):
        cell = ws.cell(4, c, h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center; cell.border = border
        ws.column_dimensions[cell.column_letter].width = w
        if ayuda:
            cell.comment = Comment(ayuda, "SciBack"); cell.comment.width = 280; cell.comment.height = 120
        # fila ejemplo (gris, row 5)
        ex = ws.cell(5, c, ej); ex.font = Font(italic=True, color=GRIS, size=9); ex.alignment = left; ex.border = border
        # filas vacías con formato
        for r in range(6, 26):
            v = ws.cell(r, c); v.fill = oblig_fill if oblig else opt_fill; v.border = border; v.alignment = left
    ws.cell(5, 1).value = "← FILA DE EJEMPLO (borrar antes de entregar)"
    ws.cell(5, 1).font = Font(italic=True, bold=True, color="C00000", size=9)
    ws.freeze_panes = "A5"
    return ws

def lista(ws, col_letter, opciones, first=6, last=200):
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(opciones), allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"{col_letter}{first}:{col_letter}{last}")

# ── Hoja Instrucciones ───────────────────────────────────────────────────────
ws0 = wb.active; ws0.title = "LÉAME"; ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 110
txt = [
    ("Plantilla de datos — Implementación sistema escolar (SciBack Odoo)", 14, True, AZUL),
    ("", 10, False, "000000"),
    ("Cómo llenar este archivo:", 11, True, "000000"),
    ("1. Cada pestaña de abajo es una lista distinta. Llene una fila por registro.", 10, False, "000000"),
    ("2. Las columnas en NARANJA son obligatorias. Las AZULES son deseables pero opcionales.", 10, False, "000000"),
    ("3. La fila gris de ejemplo muestra el formato esperado — BÓRRELA antes de entregar.", 10, False, "000000"),
    ("4. Pase el mouse sobre el encabezado (triángulo rojo) para ver la ayuda de cada campo.", 10, False, "000000"),
    ("5. Fechas en formato DD/MM/AAAA. Montos en soles, solo números (ej: 250.00).", 10, False, "000000"),
    ("6. No cambie los nombres de las pestañas ni de los encabezados.", 10, False, "000000"),
    ("", 10, False, "000000"),
    ("Pestañas:", 11, True, "000000"),
    ("• Colegio      → datos institucionales (1 sola fila)", 10, False, "000000"),
    ("• Estudiantes  → nómina completa de alumnos matriculados 2026", 10, False, "000000"),
    ("• Apoderados   → padre/madre/tutor que paga y recibe la boleta electrónica", 10, False, "000000"),
    ("• Docentes     → personal docente por nivel", 10, False, "000000"),
    ("• Pensiones    → estructura económica por nivel (matrícula, pensión, cuotas)", 10, False, "000000"),
    ("", 10, False, "000000"),
    ("Importante (privacidad — Ley 29733):", 11, True, "C00000"),
    ("Estos datos incluyen información de menores de edad. Envíelos por un canal seguro", 10, False, "000000"),
    ("(no por WhatsApp/correo personal). Se usarán solo para configurar el sistema del colegio.", 10, False, "000000"),
]
for i, (t, sz, b, color) in enumerate(txt, start=1):
    c = ws0.cell(i, 1, t); c.font = Font(size=sz, bold=b, color=color); c.alignment = left

# ── Hoja Colegio ─────────────────────────────────────────────────────────────
wc = hoja("Colegio", "Datos del colegio (una sola fila)", [
    ("Nombre / Razón social SUNAT", True, 32, "Nombre legal exacto como figura en RUC", "I.E.P. Agua Viva"),
    ("RUC", True, 14, "11 dígitos", "20512345678"),
    ("Dirección fiscal", True, 32, "Dirección registrada en SUNAT", "Av. Los Olivos 123, Lima"),
    ("Niveles que ofrece", True, 26, "Inicial / Primaria / Secundaria (los que apliquen)", "Inicial, Primaria, Secundaria"),
    ("Cód. modular Inicial", False, 16, "7 dígitos SIAGIE del nivel Inicial", "0512345"),
    ("Cód. modular Primaria", False, 16, "7 dígitos SIAGIE del nivel Primaria", "0512346"),
    ("Cód. modular Secundaria", False, 16, "7 dígitos SIAGIE del nivel Secundaria", "0512347"),
    ("UGEL", True, 18, "UGEL de jurisdicción", "UGEL 02"),
    ("Año lectivo", True, 12, "Año escolar", "2026"),
    ("Director(a) - nombre completo", True, 28, "Apellidos y nombres", "Pérez Quispe, Juan"),
    ("Director(a) - DNI", True, 12, "DNI del director", "44556677"),
    ("Email institucional", True, 26, "Correo oficial del colegio", "informes@aguaviva.edu.pe"),
    ("Teléfono institucional", True, 16, "Fijo o celular", "01-4567890"),
])

# ── Hoja Estudiantes ─────────────────────────────────────────────────────────
we = hoja("Estudiantes", "Nómina de estudiantes matriculados 2026", [
    ("Código estudiante", False, 14, "Código interno del colegio si existe", "E-0001"),
    ("Apellido paterno", True, 18, "", "Ramírez"),
    ("Apellido materno", True, 18, "", "Torres"),
    ("Nombres", True, 22, "Uno o más nombres", "María Fernanda"),
    ("Tipo documento", True, 14, "DNI / CE / Pasaporte", "DNI"),
    ("Número documento", True, 16, "DNI del alumno (8 díg). Clave para SIAGIE", "78451236"),
    ("Fecha nacimiento", True, 14, "DD/MM/AAAA", "14/03/2015"),
    ("Sexo", True, 8, "M / F", "F"),
    ("Nivel", True, 14, "Inicial / Primaria / Secundaria", "Primaria"),
    ("Grado", True, 16, "Ej: 3 años, 1° Primaria, 5° Secundaria", "5° Primaria"),
    ("Sección", True, 10, "A / B / C / Única", "A"),
    ("Condición", False, 14, "Nuevo / Continuador", "Continuador"),
    ("Nacionalidad", False, 14, "", "Peruana"),
    ("Grupo sanguíneo", False, 12, "Opcional (ficha médica)", "O+"),
    ("Dirección domicilio", False, 28, "", "Jr. Las Flores 456, Lima"),
    ("Doc. apoderado que paga", True, 18, "Nº documento del apoderado (debe existir en pestaña Apoderados)", "40556612"),
])
lista(we, "E", ["DNI", "CE", "Pasaporte"])
lista(we, "H", ["M", "F"])
lista(we, "I", ["Inicial", "Primaria", "Secundaria"])
lista(we, "K", ["A", "B", "C", "D", "Única"])
lista(we, "L", ["Nuevo", "Continuador"])

# ── Hoja Apoderados ──────────────────────────────────────────────────────────
wa = hoja("Apoderados", "Apoderados / responsables de pago (reciben boleta electrónica)", [
    ("Tipo documento", True, 14, "DNI / RUC / CE", "DNI"),
    ("Número documento", True, 16, "Identifica al apoderado; lo referencia el alumno", "40556612"),
    ("Apellidos y nombres / Razón social", True, 34, "Si es empresa que paga, su razón social", "Ramírez Soto, Carlos"),
    ("Parentesco", True, 14, "Padre / Madre / Tutor / Empresa", "Padre"),
    ("Email", True, 28, "OBLIGATORIO: a este correo llega la boleta SUNAT", "carlos.ramirez@gmail.com"),
    ("Teléfono / Celular", True, 16, "", "987654321"),
    ("Dirección", False, 30, "Para el comprobante", "Jr. Las Flores 456, Lima"),
])
lista(wa, "A", ["DNI", "RUC", "CE"])
lista(wa, "D", ["Padre", "Madre", "Tutor", "Empresa"])

# ── Hoja Docentes ────────────────────────────────────────────────────────────
wd = hoja("Docentes", "Personal docente", [
    ("Apellido paterno", True, 18, "", "Gonzales"),
    ("Apellido materno", True, 18, "", "Mamani"),
    ("Nombres", True, 22, "", "Rosa Elena"),
    ("Tipo documento", True, 14, "DNI / CE", "DNI"),
    ("Número documento", True, 16, "", "41223344"),
    ("Fecha nacimiento", True, 14, "DD/MM/AAAA (requerido por el sistema)", "22/07/1985"),
    ("Sexo", True, 8, "M / F", "F"),
    ("Email", False, 26, "", "rosa.gonzales@aguaviva.edu.pe"),
    ("Celular", False, 14, "", "999111222"),
    ("Nivel(es) que enseña", False, 20, "Inicial / Primaria / Secundaria", "Primaria"),
    ("Área / curso principal", False, 22, "Ej: Matemática, Tutoría", "Matemática"),
])
lista(wd, "D", ["DNI", "CE"])
lista(wd, "G", ["M", "F"])

# ── Hoja Pensiones ───────────────────────────────────────────────────────────
wp = hoja("Pensiones", "Estructura económica por nivel (matrícula y pensión)", [
    ("Nivel", True, 16, "Inicial / Primaria / Secundaria", "Primaria"),
    ("Monto matrícula (S/)", True, 18, "Pago único anual", "300.00"),
    ("Monto pensión mensual (S/)", True, 22, "Cuota mensual", "250.00"),
    ("N° de cuotas", True, 12, "Normalmente 10 (marzo–diciembre)", "10"),
    ("Mes inicio", True, 12, "Primer mes de cobro", "Marzo"),
    ("Mes fin", True, 12, "Último mes de cobro", "Diciembre"),
    ("Día de vencimiento", True, 16, "Día del mes en que vence la pensión", "5"),
    ("Descuento pronto pago (%)", False, 20, "Opcional", "5"),
    ("Descuento por hermano (%)", False, 20, "Opcional", "10"),
])
lista(wp, "A", ["Inicial", "Primaria", "Secundaria"])
lista(wp, "E", ["Marzo","Abril","Mayo","Junio","Julio","Agosto","Setiembre","Octubre","Noviembre","Diciembre"])
lista(wp, "F", ["Marzo","Abril","Mayo","Junio","Julio","Agosto","Setiembre","Octubre","Noviembre","Diciembre"])

wb.save("/Users/alberto/proyectos/sciback/sciback-odoo/docs/plantilla-datos-colegio.xlsx")
print("OK -> docs/plantilla-datos-colegio.xlsx")
